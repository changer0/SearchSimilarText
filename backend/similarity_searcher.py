# backend/similarity_searcher.py

# 导入所需的库
import pdfplumber  # 用于从 PDF 中提取文本
import pandas as pd  # 用于处理 Excel 文件
import re  # 用于正则表达式操作
from sentence_transformers import SentenceTransformer  # 用于生成句子嵌入
import faiss  # 用于高效的向量检索
import numpy as np  # 用于数值计算
import logging  # 用于日志记录
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment


class SimilaritySearcher:
    def __init__(self, source_pdf_path, query_excel_path=None, queries=None, output_path=None,
                 query_sheet_name='Sheet1', query_column='Query',
                 split_max_length=500, model_name='BAAI/bge-large-zh-v1.5',
                 top_k=3, threshold=0.5):
        """
        初始化相似度检索器。

        参数:
        - source_pdf_path (str): 来源 PDF 文件路径。
        - query_excel_path (str, optional): 查询 Excel 文件路径。如果提供，将忽略 `queries` 参数。
        - queries (list of str, optional): 直接提供的查询列表。如果提供，将忽略 `query_excel_path`。
        - output_path (str, optional): 输出 Excel 文件路径。如果不提供，将不生成 Excel 文件。
        - query_sheet_name (str): Excel 中查询条目所在的工作表名称。
        - query_column (str): Excel 中查询条目所在的列名。
        - split_max_length (int): 文本分段的最大长度。
        - model_name (str): 用于生成嵌入的预训练模型名称。
        - top_k (int): 每个查询检索的相关段落数量。
        - threshold (float): 判断是否相关的相似度阈值。
        """
        self.source_pdf_path = source_pdf_path
        self.query_excel_path = query_excel_path
        self.output_path = output_path
        self.query_sheet_name = query_sheet_name
        self.query_column = query_column
        self.split_max_length = split_max_length
        self.model_name = model_name
        self.top_k = top_k
        self.threshold = threshold

        # 初始化其他属性
        self.file_a_text = ""
        self.file_a_segments = []
        self.file_a_embeddings = None
        self.embedding_model = None
        self.faiss_index = None
        self.queries = queries if queries is not None else []
        self.df_output = None

    def extract_text_from_pdf(self):
        """从 PDF 文件中提取文本内容。"""
        text = ""
        try:
            with pdfplumber.open(self.source_pdf_path) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n"
            self.file_a_text = text
            logging.info("成功提取 PDF 文件中的文本内容。")
        except Exception as e:
            logging.error(f"提取 PDF 内容时出错: {e}")
            raise e

    def read_queries(self):
        """读取 Excel 文件中的查询条目，或使用直接提供的查询列表。"""
        if self.queries:
            logging.info(f"使用直接提供的 {len(self.queries)} 个查询条目。")
            return
        try:
            df = pd.read_excel(self.query_excel_path, sheet_name=self.query_sheet_name)
            self.queries = df[self.query_column].dropna().tolist()
            logging.info(f"成功读取 Excel 文件中的 {len(self.queries)} 个查询条目。")
        except Exception as e:
            logging.error(f"读取查询条目时出错: {e}")
            raise e

    def split_text(self):
        """将文本按指定的最大长度分段。"""
        sentences = re.split(r'(?<=[。！？；])', self.file_a_text)
        current_segment = ""
        for sentence in sentences:
            if len(current_segment) + len(sentence) <= self.split_max_length:
                current_segment += sentence
            else:
                self.file_a_segments.append(current_segment)
                current_segment = sentence
        if current_segment:
            self.file_a_segments.append(current_segment)
        logging.info(f"成功将文本分段为 {len(self.file_a_segments)} 个段落。")

    def normalize_embeddings(self, embeddings):
        """
        将嵌入向量归一化为单位向量。

        参数:
        - embeddings (np.ndarray): 原始嵌入向量。

        返回:
        - normalized_embeddings (np.ndarray): 归一化后的嵌入向量。
        """
        norms = np.linalg.norm(embeddings, axis=1, keepdims=True)
        norms[norms == 0] = 1  # 防止除以零
        normalized_embeddings = embeddings / norms
        logging.info("成功归一化嵌入向量。")
        return normalized_embeddings

    def generate_embeddings(self):
        """生成文本段落的嵌入向量。"""
        try:
            self.file_a_embeddings, self.embedding_model = self._generate_embeddings_internal()
            logging.info("成功生成嵌入向量。")
        except Exception as e:
            logging.error(f"生成嵌入向量时出错: {e}")
            raise e

    def _generate_embeddings_internal(self):
        """内部方法：生成嵌入向量。"""
        model = SentenceTransformer(self.model_name)
        embeddings = model.encode(self.file_a_segments, convert_to_tensor=False, show_progress_bar=True)
        embeddings = np.array(embeddings).astype('float32')
        embeddings = self.normalize_embeddings(embeddings)
        return embeddings, model

    def build_faiss_index_cosine(self):
        """构建 FAISS 内积索引，用于余弦相似度检索。"""
        try:
            dimension = self.file_a_embeddings.shape[1]
            logging.info(f"使用 {dimension} 维度构建 FAISS 内积索引。")
            index = faiss.IndexFlatIP(dimension)  # 使用内积索引，需要注意，如果使用内积索引必须进行向量归一化
            index.add(self.file_a_embeddings)
            self.faiss_index = index
            logging.info(f"成功构建 FAISS 内积索引，包含 {index.ntotal} 个向量。")
        except Exception as e:
            logging.error(f"构建 FAISS 内积索引时出错: {e}")
            raise e

    def retrieve_cosine_similarity(self, query):
        """
        检索与查询最相关的文本段落，基于余弦相似度。

        参数:
        - query (str): 查询字符串。

        返回:
        - results (list): 包含相关段落和相似度的字典列表。
        """
        if not self.faiss_index or not self.embedding_model:
            logging.warning("FAISS 索引或嵌入模型未初始化。")
            return []

        try:
            query_embedding = self.embedding_model.encode([query], convert_to_tensor=False)
            query_embedding = np.array(query_embedding).astype('float32')
            query_embedding = self.normalize_embeddings(query_embedding)
        except Exception as e:
            logging.error(f"生成查询嵌入向量时出错: {e}")
            raise e

        try:
            """
            similarities: 一个二维数组，包含每个查询与检索到的段落之间的相似度分数。
            indices: 一个二维数组，包含每个查询对应的最相关段落在原始文本段落列表中的索引位置。 
            """
            similarities, indices = self.faiss_index.search(query_embedding, self.top_k)
        except Exception as e:
            logging.error(f"使用 FAISS 进行检索时出错: {e}")
            raise e

        results = []
        # 同时迭代similarities 和 indices
        for similarity, idx in zip(similarities[0], indices[0]):
            # logging.info(f"相似度: {similarity}, 段落索引: {idx}")
            if idx < len(self.file_a_segments):
                # 清洗段落内容
                clean_paragraph = self.remove_illegal_characters(self.file_a_segments[idx])
                results.append({
                    'paragraph': clean_paragraph,
                    'similarity': round(float(similarity), 4)  # 保留4位小数
                })
        return results

    def remove_illegal_characters(self, text):
        """
        移除 Excel 不允许的非法字符。

        参数:
        - text (str): 输入的字符串。

        返回:
        - clean_text (str): 清洗后的字符串。
        """
        if not isinstance(text, str):
            return text
        illegal_chars = re.compile(r'[\x00-\x08\x0B\x0C\x0E-\x1F]')
        clean_text = illegal_chars.sub('', text)
        return clean_text

    def check_relevance_and_output_cosine(self):
        """
        检查每个查询条目在 PDF 中的相关性，并生成输出。

        返回:
        - result_json (list): 包含每个查询的相关性和相关段落信息的列表。
        """
        if not self.file_a_segments:
            logging.warning("没有可用于检索的文本段落。所有查询将标记为无相关性。")
            result_json = []
            for query in self.queries:
                clean_query = self.remove_illegal_characters(query)
                result_json.append({
                    'query': clean_query,
                    'relevance': False,
                    'relevant_paragraphs': []
                })
            return result_json

        result_json = []

        for query in self.queries:
            logging.info(f"正在处理查询: {query}")
            clean_query = self.remove_illegal_characters(query)
            # 检索当前查询最相关的段落
            results = self.retrieve_cosine_similarity(query)
            # 判断是否有任何一个检索结果的相似度超过阈值
            has_relevance = any(res['similarity'] >= self.threshold for res in results)

            # 收集相关段落
            relevant_paragraphs = [
                {
                    'paragraph': res['paragraph'],
                    'similarity': res['similarity']
                }
                for res in results if res['similarity'] >= self.threshold
            ]

            result_json.append({
                'query': clean_query,
                'relevance': has_relevance,
                'relevant_paragraphs': relevant_paragraphs
            })

        logging.info("完成所有查询的相关性检索和数据准备。")
        return result_json

    def save_output_excel(self):
        """保存结果到新的 Excel 文件，并优化格式（如果指定了 output_path）。"""
        if self.output_path is None:
            logging.info("未指定 output_path，不保存 Excel 文件。")
            return
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "相关性结果"

            # 定义表头
            headers = ['Query', 'Relevance', 'Relevant Paragraphs', 'Similarities']
            ws.append(headers)

            # 定义边框样式
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for record in self.df_output:
                query = record['query']
                relevance = "Yes" if record['relevance'] else "No"
                if record['relevance']:
                    paragraphs = "\n\n".join(
                        [f"Paragraph {i + 1}: {p['paragraph']}" for i, p in enumerate(record['relevant_paragraphs'])])
                    similarities = "\n\n".join(
                        [f"Paragraph {i + 1} Similarity: {p['similarity']}" for i, p in enumerate(record['relevant_paragraphs'])])
                else:
                    paragraphs = ""
                    similarities = ""
                ws.append([query, relevance, paragraphs, similarities])

            # 应用边框和调整列宽
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=4):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(wrap_text=True, vertical='top')

            # 调整列宽
            column_widths = {
                'A': 30,  # Query
                'B': 10,  # Relevance
                'C': 100,  # Relevant Paragraphs
                'D': 20  # Similarities
            }
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width

            wb.save(self.output_path)
            logging.info(f"成功保存 Excel 结果到 {self.output_path}")
        except Exception as e:
            logging.error(f"保存 Excel 文件时出错: {e}")
            raise e

    def run(self):
        """执行完整的流程，并返回 JSON 数据。"""
        # 提取文本
        self.extract_text_from_pdf()
        # 分割文本
        self.split_text()
        # 读取查询
        self.read_queries()
        # 如果没有段落，跳过嵌入生成和检索步骤
        if not self.file_a_segments:
            self.df_output = self.check_relevance_and_output_cosine()
        else:
            # 生成嵌入向量
            self.generate_embeddings()
            # 构建 FAISS 索引
            self.build_faiss_index_cosine()
            # 检索相关性并输出
            self.df_output = self.check_relevance_and_output_cosine()
        # 保存结果到 Excel（如果指定了 output_path）
        self.save_output_excel()
        # 返回 JSON 数据
        logging.info("完成所有步骤，处理成功。")
        return self.df_output
